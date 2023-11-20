import click
import requests
import json
import re
import openpyxl
import csv
import xlrd
from datetime import datetime
from urllib.parse import urlparse, urljoin, parse_qs

def datetime_to_date_str(value: datetime) -> str:
    return value.strftime('%d/%m/%Y')

def str_to_bool(value: str) -> bool:
    if value is not None:
        if value in ['1', 'yes', 'NAI', 'Ναι', 'true', 'True']:
            return True
        else:
            return False
    else:
        return False

def filter_cvs_column(value: str) -> str:
    """
    Filters out value like '=""123""'
    # https://regex101.com/r/ASMSuj/3
    """
    if value is None:
        return value

    g = re.match(r'\"=\"\"(.*)\"\"\"', value)    
    
    if g is not None:
        return g.group(1)
    else:
        return value


def string_or_null(value: str) -> str:
    
    if value is None:
        return None

    if len(value.strip()) > 0:
        return value
    else:
        return None 


def is_empty_or_null(value: str) -> bool:
    return value is None or len(value) == 0


@click.group()
@click.option('--debug', default=False, is_flag=True)
@click.option('--phaistos_api', default='http://localhost:8000')
@click.pass_context
def cli(ctx, debug, phaistos_api):
    # ensure that ctx.obj exists and is a dict (in case `cli()` is called
    # by means other than the `if` block below)
    ctx.ensure_object(dict)

    ctx.obj['debug'] = debug
    ctx.obj['phaistos_api'] = phaistos_api


@cli.command()
@click.argument('report_04_01_path', type=click.File('r', encoding='cp1253'))
@click.option('--employee_am', default=None, type=str, help='AM of employee')
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--employee_type', default=None, type=click.Choice(['Μόνιμος', 'Αναπληρωτής', 'Αναπληρωτής ΠΔΕ']), help='employee type')
@click.option('--skip_until_am', default=None, type=int, help='skip until employee AM')
@click.option('--skip_no_current_unit', default=False, is_flag=True, help='skip employee if no current unit is set')
@click.option('--continue_after_am', default=None, type=int, help='continue after employee AM')
@click.pass_context
def import_employee_report_04_01(ctx, report_04_01_path, employee_am, employee_afm, employee_type, skip_until_am, 
                                 continue_after_am, skip_no_current_unit):
    """Import myschool employee report 01 from REPORT_04_01_PATH
    
    """
    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    employee_resource = phaistos_api + "/api/bulk_import/myschool/employees/"
    
    employee_reader = csv.reader(report_04_01_path, delimiter=';', quotechar='|')
    row1 = next(employee_reader)  # gets the first line

    with requests.Session() as s:
        
        for row in employee_reader:

            _employee_am = row[0]
            
            if employee_am is not None and employee_am != _employee_am:
                continue
            
            _employee_afm = filter_cvs_column(row[1])
            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue
            
            _employee_sex = row[2]
            _employee_last_name = row[3]
            _employee_first_name = row[4]
            _employee_father_name = row[5]
            _employee_mother_name = row[6]
            _employee_birthday = row[51]
            _employee_telephone = row[9]
            _employee_mobile = row[10]
            _employee_email = row[12]
            _employee_email_psd = row[13]
            _employee_type_name = row[47]

            if employee_type is not None and employee_type != _employee_type_name:
                continue

            _employee_current_unit_id = row[35]
            _employee_current_unit_name = row[36]
            _employee_specialization_id = row[14]
            _employee_specialization_name = row[15]
            _employee_mandatory_week_workhours = row[25]
            _employee_mk = row[19]
            _employee_bathmos = row[18]
            _employee_first_workday_date = row[32]
            _employee_fek_diorismou = row[20]
            _employee_fek_diorismou_date = row[21]
    
            # normalization
            _employee_current_unit_id = filter_cvs_column(_employee_current_unit_id)

            employee_dict = {
                'employee_am': _employee_am,
                'employee_afm': _employee_afm,
                'employee_sex': _employee_sex,
                'employee_last_name': _employee_last_name,
                'employee_first_name': _employee_first_name,
                'employee_father_name': _employee_father_name,
                'employee_mother_name': _employee_mother_name,
                'employee_telephone': _employee_telephone,
                'employee_mobile': _employee_mobile,
                'employee_email': _employee_email,
                'employee_email_psd': _employee_email_psd,
                'employee_type_name': _employee_type_name,
                'employee_current_unit_id': _employee_current_unit_id,
                'employee_current_unit_name': _employee_current_unit_name,
                'employee_specialization_id': _employee_specialization_id,
                'employee_specialization_name': _employee_specialization_name,
                'employee_mandatory_week_workhours': _employee_mandatory_week_workhours,
                'employee_mk': _employee_mk,
                'employee_bathmos': _employee_bathmos,
                'employee_first_workday_date': _employee_first_workday_date,
                'employee_fek_diorismou': _employee_fek_diorismou,
                'employee_fek_diorismou_date': _employee_fek_diorismou_date,
                'employee_birthday': _employee_birthday
            }

            employee_label = f"({employee_dict.get('employee_am')}) {employee_dict.get('employee_last_name')} {employee_dict.get('employee_first_name')} {employee_dict.get('employee_father_name')} [{employee_dict.get('employee_type_name')}]"

            if is_empty_or_null(_employee_current_unit_id):
                if skip_no_current_unit:
                    click.echo(f"[W] skipping '{employee_label}' since it has no current unit")
                    continue
                else:
                    employee_dict['employee_current_unit_id'] = '319'
                    employee_dict['employee_current_unit_name'] = 'Δ/ΝΣΗ Β/ΜΙΑΣ ΕΚΠ/ΣΗΣ Ν. ΗΡΑΚΛΕΙΟΥ'
        

            if debug:
                click.echo(f"[I] request object is {json.dumps(employee_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            try:
                r = s.post(employee_resource, json=employee_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added employee '{employee_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                    
                elif r.status_code == 200:
                    # employee was updated
                    click.echo(f"[I] successfully UPDATED employee '{employee_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 404:
                    # employee could not matched with phaistos
                    click.echo(f"[W] could not found employee {employee_label} in phaistos")
                    raise click.Abort()
                else:
                    click.echo(f"[W] failed inserting/updating employee '{employee_label}'")
                    click.echo(f"[W] Response : HTTP/{r.status_code}")
                    click.echo()
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)

        
    

@cli.command()
@click.argument('report_01_07_path', type=click.File('r', encoding='cp1253'))
@click.option('--employee_am', default=None, type=str, help='AM of employee')
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--skip_until_am', default=None, type=int, help='skip until employee AM')
@click.option('--continue_after_am', default=None, type=int, help='continue after employee AM')
@click.pass_context
def import_employee_report_01_07(ctx, report_01_07_path, employee_am, employee_afm, skip_until_am, continue_after_am):
    """Import myschool employee report 01 from REPORT_07_01_PATH
    
    """
    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    employee_resource = phaistos_api + "/api/bulk_import/myschool/employees/"
    
    employee_reader = csv.reader(report_01_07_path, delimiter=';', quotechar='|')
    row1 = next(employee_reader)  # gets the first line

    with requests.Session() as s:
        
        for row in employee_reader:

            _employee_am = row[0]
            
            if employee_am is not None and employee_am != _employee_am:
                continue
            
            _employee_afm = filter_cvs_column(row[1])
            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue
            
            _employee_sex = row[2]
            _employee_last_name = row[3]
            _employee_first_name = row[4]
            _employee_father_name = row[5]
            _employee_mother_name = row[6]
            _employee_birthday = row[49]
            _employee_telephone = row[9]
            _employee_mobile = row[10]
            _employee_email = row[12]
            _employee_email_psd = row[13]
            _employee_type_name = f'Διοικητικός {row[47]}'

            if _employee_type_name == 'Διοικητικός Μόνιμος':
                _employee_type_name = 'Διοικητικός'

            _employee_current_unit_id = row[35]
            _employee_current_unit_name = row[36]
            _employee_specialization_id = row[14]
            _employee_specialization_name = row[15]
            _employee_mandatory_week_workhours = row[25]
            _employee_mk = row[19]
            _employee_bathmos = row[18]
            _employee_first_workday_date = row[32]
            _employee_fek_diorismou = row[20]
            _employee_fek_diorismou_date = row[21]
    

            # normalization
            _employee_current_unit_id = filter_cvs_column(_employee_current_unit_id)

            employee_dict = {
                'employee_am': _employee_am,
                'employee_afm': _employee_afm,
                'employee_sex': _employee_sex,
                'employee_last_name': _employee_last_name,
                'employee_first_name': _employee_first_name,
                'employee_father_name': _employee_father_name,
                'employee_mother_name': _employee_mother_name,
                'employee_telephone': _employee_telephone,
                'employee_mobile': _employee_mobile,
                'employee_email': _employee_email,
                'employee_email_psd': _employee_email_psd,
                'employee_type_name': _employee_type_name,
                'employee_current_unit_id': _employee_current_unit_id,
                'employee_current_unit_name': _employee_current_unit_name,
                'employee_specialization_id': _employee_specialization_id,
                'employee_specialization_name': _employee_specialization_name,
                'employee_mandatory_week_workhours': _employee_mandatory_week_workhours,
                'employee_mk': _employee_mk,
                'employee_bathmos': _employee_bathmos,
                'employee_first_workday_date': _employee_first_workday_date,
                'employee_fek_diorismou': _employee_fek_diorismou,
                'employee_fek_diorismou_date': _employee_fek_diorismou_date,
                'employee_birthday': _employee_birthday
            }

            employee_label = f"({employee_dict.get('employee_am')}) {employee_dict.get('employee_last_name')} {employee_dict.get('employee_first_name')} {employee_dict.get('employee_father_name')} [{employee_dict.get('employee_type_name')}]"


            if debug:
                click.echo(f"[I] request object is {json.dumps(employee_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            try:
                r = s.post(employee_resource, json=employee_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added employee '{employee_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                    
                elif r.status_code == 200:
                    # employee was updated
                    click.echo(f"[I] successfully UPDATED employee '{employee_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 404:
                    # employee could not matched with phaistos
                    click.echo(f"[W] could not found employee {employee_label} in phaistos")
                    raise click.Abort()
                else:
                    click.echo(f"[W] failed inserting/updating employee '{employee_label}'")
                    click.echo(f"[W] Response : HTTP/{r.status_code}")
                    click.echo()
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)
            


@cli.command()
@click.argument('employments_report_path', type=click.Path(exists=True))
@click.option('--employee_am', default=None, type=str, help='AM of employee')
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--skip_until_am', default=None, type=int, help='skip until employee AM')
@click.option('--continue_after_am', default=None, type=int, help='continue after employee AM')
@click.pass_context
def import_employments_report(ctx, employments_report_path, employee_am, employee_afm, skip_until_am, continue_after_am):
    """Import myschool employee report 01 from REPORT_07_01_PATH

    
    """
    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    employment_resource = phaistos_api + "/api/bulk_import/myschool/employments/"
    
    book = xlrd.open_workbook(employments_report_path, encoding_override='cp1253')
    sh = book.sheet_by_index(0)
    
    with requests.Session() as s:
        
        for rx in range(2, sh.nrows):
            
            row = sh.row(rx)
            
            _employee_am = row[0].value
            
            if employee_am is not None and employee_am != _employee_am:
                continue
            
            _employee_afm = row[1].value
            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue
            
            _employee_last_name = row[2].value
            _employee_first_name = row[3].value
            _employee_specialization_id = row[4].value
            _employee_type = row[5].value
            _employee_employment_type = row[6].value
            _employee_employment_unit_id = row[7].value
            _employee_employment_unit_name = row[8].value

            # compute / parse working days
            working_days = ''
            for col in [row[9], row[10], row[11], row[12], row[13]]:
                try:
                    working_days += f'{int(col.value)}:'
                except:
                    working_days += ''
            
            if working_days.endswith(':'):
                working_days = working_days[:-1]

            try:
                _employee_employment_hours = int(row[14].value)
            except:
                _employee_employment_hours = 0
            
            _employee_employment_from = row[15].value
            _employee_employment_from = datetime(*xlrd.xldate_as_tuple(_employee_employment_from, book.datemode))
            
            _employee_employment_until = row[16].value
            _employee_employment_until = datetime(*xlrd.xldate_as_tuple(_employee_employment_until, book.datemode))

            _employee_employment_status = row[17].value

            # _employee_email = row[12]
            # _employee_email_psd = row[13]
            # _employee_type_name = f'Διοικητικός {row[47]}'

            # if _employee_type_name == 'Διοικητικός Μόνιμος':
            #     _employee_type_name = 'Διοικητικός'

            # _employee_current_unit_id = row[35]
            # _employee_current_unit_name = row[36]
            # _employee_specialization_id = row[14]
            # _employee_specialization_name = row[15]
            # _employee_mandatory_week_workhours = row[25]
            # _employee_mk = row[19]
            # _employee_bathmos = row[18]
            # _employee_first_workday_date = row[32]
            # _employee_fek_diorismou = row[20]
            # _employee_fek_diorismou_date = row[21]
    

            
            employee_dict = {
                'employee_am': _employee_am,
                'employee_afm': _employee_afm,
                'employee_last_name': _employee_last_name,
                'employee_first_name': _employee_first_name,
                'employee_employment_unit_id': _employee_employment_unit_id,
                'employee_employment_unit_name': _employee_employment_unit_name,
                'employee_specialization_id': _employee_specialization_id,
                'employee_type': _employee_type,
                'employee_employment_type': _employee_employment_type,
                'employee_employment_days': working_days,
                'employee_employment_hours': _employee_employment_hours,
                'employee_employment_from': datetime_to_date_str(_employee_employment_from),
                'employee_employment_until': datetime_to_date_str(_employee_employment_until),
                'employee_employment_status': _employee_employment_status,
            }

            print(employee_dict)
            employment_label = f"({employee_dict.get('employee_am')}) {employee_dict.get('employee_last_name')} {employee_dict.get('employee_first_name')} {employee_dict.get('employee_father_name')} [{employee_dict.get('employee_type_name')}]"


            if debug:
                click.echo(f"[I] request object is {json.dumps(employee_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            
            
        
            try:
                r = s.post(employment_resource, json=employee_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added employment '{employment_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                    
                elif r.status_code == 200:
                    # employee was updated
                    click.echo(f"[I] successfully UPDATED employment '{employment_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 404:
                    # employee could not matched with phaistos
                    click.echo(f"[W] could not found employment {employment_label} in phaistos")
                    raise click.Abort()
                else:
                    click.echo(f"[W] failed inserting/updating employment '{employment_label}'")
                    click.echo(f"[W] Response : HTTP/{r.status_code}")
                    click.echo()
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)
            

@cli.command()
@click.argument('report_path', type=click.Path(exists=True))
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--skip_until_afm', default=None, type=int, help='skip until employee AFM')
@click.option('--continue_after_afm', default=None, type=int, help='continue after employee AFM')
@click.option('--dide', default='ΔΙΕΥΘΥΝΣΗ Δ.Ε. ΗΡΑΚΛΕΙΟΥ', help='Τοποθέτηση Δ/ΝΣΗ ΕΚΠ/ΣΗΣ')
@click.option('--phase', help='Φάση Προσλήψεων', required=True)
@click.pass_context
def import_deputy_hiring_report(ctx, report_path, employee_afm, dide, phase, skip_until_afm, continue_after_afm):
    """
    Import Deputy hiring announcement
    
    """
    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    api_resource = phaistos_api + "/api/bulk_import/substitute_employment_announcement/"
    
    book = openpyxl.load_workbook(report_path)
    sh = book.worksheets[0]
    
    with requests.Session() as s:
        for row in sh.iter_rows(min_row=2, max_row=sh.max_row):
        
            
            #row = sh.row(rx)
            _xrimatodotisi = row[0].value
            _aa = row[1].value
            _aa_rois = row[2].value
            _source = row[3].value
            _employee_afm = row[4].value

            _employee_last_name = row[5].value
            _employee_first_name = row[6].value
            _employee_father_name = row[7].value
            _employee_mother_name = row[8].value
            _employee_klados_id = row[9].value
            _employee_specialization_id = row[10].value
            _pinakas = row[11].value
            _seira_pinaka = row[12].value
            _moria_pinaka = row[13].value
            _perioxh_topothetisis = row[14].value
            _orario = row[15].value
            _dide = row[16].value
            _periferia = row[17].value
            _employee_address_city = row[18].value
            _employee_address_line = row[19].value
            _employee_address_postal_code = row[20].value
            _employee_telephone = row[21].value
            _employee_mobile = row[22].value
            _employee_email = row[23].value
            _employee_birthday = row[24].value
            _employee_adt = row[25].value
            _proslipsi = row[26].value


            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue

            if dide != _dide:
                continue
            
            request_dict = {
                'phase': phase,
                'employee_afm': _employee_afm,
                'employee_last_name': _employee_last_name,
                'employee_first_name': _employee_first_name,
                'employee_father_name':_employee_father_name,
                'employee_mother_name': _employee_mother_name,
                'employee_klados_id' :_employee_klados_id,
                'employee_specialization_id': _employee_specialization_id,
                'financing_source_code': _xrimatodotisi,
                'employment_source_code': _source,
                'employment_table': _pinakas,
                'employment_table_position': _seira_pinaka,
                'employment_table_score': _moria_pinaka,
                'employment_workhour_type': _orario, 
                'employee_address_city': _employee_address_city,
                'employee_address_line': _employee_address_line,
                'employee_address_postal_code': _employee_address_postal_code,
                'employee_telephone': _employee_telephone,
                'employee_mobile': _employee_mobile,
                'employee_email': _employee_email,
                'employee_birthday': datetime_to_date_str(_employee_birthday),
                'employee_adt': _employee_adt,
                
            }

            
            employment_label = f"({request_dict.get('employee_am')}) {request_dict.get('employee_last_name')} {request_dict.get('employee_first_name')} {request_dict.get('employee_father_name')} [{request_dict.get('employee_type_name')}]"


            if debug:
                click.echo(f"[I] request object is {json.dumps(request_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            
            
    
            try:
                r = s.post(api_resource, json=request_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added employment '{employment_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                    
                elif r.status_code == 200:
                    # employee was updated
                    click.echo(f"[I] successfully UPDATED employment '{employment_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 404:
                    # employee could not matched with phaistos
                    click.echo(f"[W] could not found employment {employment_label} in phaistos")
                    raise click.Abort()
                else:
                    click.echo(f"[W] failed inserting/updating employment '{employment_label}'")
                    click.echo(f"[W] Response : HTTP/{r.status_code}")
                    click.echo()
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)
            

@cli.command()
@click.argument('report_path', type=click.Path(exists=True))
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--skip_until_afm', default=None, type=int, help='skip until employee AFM')
@click.option('--continue_after_afm', default=None, type=int, help='continue after employee AFM')
@click.option('--phase', help='Φάση Προσλήψεων', required=True)
@click.pass_context
def import_deputy_placement_report(ctx, report_path, employee_afm, phase, skip_until_afm, continue_after_afm):
    """
    Import Deputy placement announcement (Απόφαση Τοποθέτησης Αναπληρωτών)
    
    """

    # phaistos_importer --debug import-deputy-placement-report "ΓΕΝΙΚΗΣ ΠΔΕ ΠΕΡΙΣΥΝΟ.xlsx" --phase="Lala"

    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    api_resource = phaistos_api + "/api/bulk_import/substitute_employment_placement/"
    
    book = openpyxl.load_workbook(report_path)
    sh = book.worksheets[0]

    
    # determine indexes
    header_row = sh[1]
    for cell in header_row:
        cell_value = cell.value
        col_idx = cell.col_idx - 1
        if cell_value in ['ΑΦΜ', 'Α.Φ.Μ.']:
            _employee_afm_idx = col_idx
        elif cell_value in ['ΗΜ. ΠΡΟΣΛΗΨΗΣ']:
            _employment_start_date_idx= col_idx
        elif cell_value in ['ΕΠΙΘΕΤΟ']:
            _employee_last_name_idx = col_idx
        elif cell_value in ['ΟΝΟΜΑ']:
            _employee_first_name_idx = col_idx
        elif cell_value in ['ΕΙΔΙΚΟΤΗΤΑ']:
            _employement_specialization_idx = col_idx
        elif cell_value in ['ΩΡΑΡΙΟ']:
            _employment_hour_type_idx = col_idx
        elif cell_value in ['ΚΩΔ. ΣΧΟΛΕΙΟΥ']:
            _employement_school_code_idx = col_idx
        elif cell_value in ['ΣΧ. ΑΝΑΛΗΨΗΣ']:
            _employement_is_main_school_idx = col_idx
        elif cell_value in ['ΩΡΕΣ']:
            _employment_work_hours_idx = col_idx
        elif cell_value in ['ΤΥΠΟΣ ΚΕΝΟΥ']:
            _employment_source_code_idx = col_idx
    
    with requests.Session() as s:
        for row in sh.iter_rows(min_row=2, max_row=sh.max_row):
            
            #row = sh.row(rx)
            _employment_start_date = datetime_to_date_str(row[_employment_start_date_idx].value)
            _employee_afm = row[_employee_afm_idx].value
            _employee_last_name = row[_employee_last_name_idx].value
            _employee_first_name = row[_employee_first_name_idx].value
            _employement_specialization = row[_employement_specialization_idx].value
            _employment_hour_type = row[_employment_hour_type_idx].value
            _employment_work_hours = row[_employment_work_hours_idx].value
            _employement_school_code = row[_employement_school_code_idx].value
            _employement_is_main_school = row[_employement_is_main_school_idx].value 
            _employment_source_code = row[_employment_source_code_idx].value
            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue
            
            request_dict = {
                'phase': phase,
                'employment_start_date': _employment_start_date,
                'employee_afm': _employee_afm,
                'employee_last_name': _employee_last_name,
                'employee_first_name': _employee_first_name,
                'employement_specialization_id': _employement_specialization,
                'employment_source_code': _employment_source_code,
                'employment_hour_type': _employment_hour_type,
                'employment_work_hours': _employment_work_hours,
                'employement_school_code': _employement_school_code,
                'employement_is_main_school':  str_to_bool(_employement_is_main_school)
            }

            print(_employement_is_main_school)

             
            employment_label = f"({request_dict.get('employee_am')}) {request_dict.get('employee_last_name')} {request_dict.get('employee_first_name')} {request_dict.get('employee_father_name')} [{request_dict.get('employee_type_name')}]"


            if debug:
                click.echo(f"[I] request object is {json.dumps(request_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            
            try:
                r = s.post(api_resource, json=request_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added employment '{employment_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 200:
                    click.echo(f"[I] employment alreay found '{employment_label}' with ID {r.json().get('id')}")
                elif r.status_code == 404:
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    click.echo(f"[W] could not found hiring announcement for placement '{employment_label}'")
                    continue
                else:
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    click.echo(f"[W] {r.status_code} : could to process {employment_label} in phaistos")
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)
            

@cli.command()
@click.argument('report_path', type=click.File('r', encoding='cp1253'))
@click.option('--employee_afm', default=None, type=str, help='AFM of employee')
@click.option('--skip_until_afm', default=None, type=int, help='skip until employee AFM')
@click.option('--continue_after_afm', default=None, type=int, help='continue after employee AFM')
@click.pass_context
def import_school_principals(ctx, report_path, employee_afm, skip_until_afm, continue_after_afm):
    """
    Import School Principals (report 4.25)
    
    """

    # phaistos_importer --debug import-school-principals "stat4_25_2023-11-08-104103.csv"

    started_on = datetime.now().replace(microsecond=0)
    debug = ctx.obj.get('debug', False)
    phaistos_api = ctx.obj['phaistos_api']
    api_resource = phaistos_api + "/api/bulk_import/myschool/schoolprincipals/"
    
    csv_reader = csv.reader(report_path, delimiter=';', quotechar='|')
    row1 = next(csv_reader)  # gets the first line    
    
    with requests.Session() as s:
        for row in csv_reader:
            
            
            #row = sh.row(rx)
            _employee_first_name = row[18]
            _employee_last_name = row[17]
            _employee_father_name = row[19]
            _specialization_code = row[25]

            _employee_am = row[14]
            _employee_afm = filter_cvs_column(row[15])
            _assignment_unit_id = filter_cvs_column(row[7])
            _assignment_unit_name = row[8]
            
            if employee_afm is not None and employee_afm != _employee_afm:
                continue
            
            request_dict = {
                'employee_afm': _employee_afm,
                'employee_am': _employee_am,
                'employee_first_name': _employee_first_name,
                'employee_last_name': _employee_last_name,
                'employee_father_name': _employee_father_name,
                'specialization_code': _specialization_code,
                'assignment_unit_id': _assignment_unit_id,
            }

             
            school_principal_label = f"({request_dict.get('employee_am')}) {request_dict.get('employee_last_name')} {request_dict.get('employee_first_name')} {request_dict.get('employee_father_name')} [{request_dict.get('specialization_code')}]"


            if debug:
                click.echo(f"[I] request object is {json.dumps(request_dict, ensure_ascii=False, sort_keys=True, indent=2)}")
            
            try:
                r = s.post(api_resource, json=request_dict)
                
                if r.status_code == 201:
                    # employee was created
                    click.echo(f"[I] successfully added school principal '{school_principal_label}' with ID {r.json().get('id')}")
                    #click.echo(json.dumps(r.json(), sort_keys=True, indent=2))
                elif r.status_code == 200:
                    click.echo(f"[I] school principal already found '{school_principal_label}' with ID {r.json().get('id')}")
                elif r.status_code == 404:
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    click.echo(f"[W] could not add school principal '{school_principal_label}'")
                    raise click.Abort()
                else:
                    click.echo(json.dumps(r.json(), sort_keys=True, ensure_ascii=False, indent=2))
                    click.echo(f"[W] {r.status_code} : could to process {school_principal_label} in phaistos")
                    raise click.Abort()
                
                
                #return True

            except Exception as e:
                raise click.ClickException(e)